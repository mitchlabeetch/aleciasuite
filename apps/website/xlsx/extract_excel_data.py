#!/usr/bin/env python3
"""
Extract data and formulas from Excel files for Alecia Colab Sheets implementation.
"""

import os
import json
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(XLSX_DIR, 'csv_exports')

os.makedirs(OUTPUT_DIR, exist_ok=True)

def extract_workbook_info(filepath):
    """Extract all data and formulas from a workbook."""
    filename = os.path.basename(filepath)
    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print('='*60)

    # Load workbook with data_only=False to get formulas
    wb_formulas = load_workbook(filepath, data_only=False)
    # Load again with data_only=True to get computed values
    wb_values = load_workbook(filepath, data_only=True)

    workbook_info = {
        'filename': filename,
        'sheets': []
    }

    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        print(f"\n  Sheet: {sheet_name}")
        print(f"  Dimensions: {ws_formulas.dimensions}")

        sheet_info = {
            'name': sheet_name,
            'dimensions': ws_formulas.dimensions,
            'max_row': ws_formulas.max_row,
            'max_col': ws_formulas.max_column,
            'formulas': [],
            'headers': [],
            'data_types': {},
            'merged_cells': [str(m) for m in ws_formulas.merged_cells.ranges]
        }

        # Extract headers (first row)
        for col in range(1, ws_formulas.max_column + 1):
            cell = ws_formulas.cell(row=1, column=col)
            if cell.value:
                sheet_info['headers'].append({
                    'column': get_column_letter(col),
                    'value': str(cell.value)
                })

        # Extract formulas
        formula_count = 0
        for row in range(1, ws_formulas.max_row + 1):
            for col in range(1, ws_formulas.max_column + 1):
                cell = ws_formulas.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(col)}{row}"
                    value_cell = ws_values.cell(row=row, column=col)
                    sheet_info['formulas'].append({
                        'cell': cell_ref,
                        'formula': cell.value,
                        'computed_value': str(value_cell.value) if value_cell.value is not None else None
                    })
                    formula_count += 1

        print(f"  Formulas found: {formula_count}")
        print(f"  Merged cells: {len(sheet_info['merged_cells'])}")

        # Export to CSV
        csv_filename = f"{os.path.splitext(filename)[0]}_{sheet_name.replace(' ', '_')}.csv"
        csv_path = os.path.join(OUTPUT_DIR, csv_filename)

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in ws_values.iter_rows(min_row=1, max_row=ws_values.max_row,
                                           min_col=1, max_col=ws_values.max_column):
                writer.writerow([cell.value for cell in row])

        print(f"  CSV exported: {csv_filename}")

        workbook_info['sheets'].append(sheet_info)

    wb_formulas.close()
    wb_values.close()

    return workbook_info

def main():
    all_workbooks = []

    # Process all Excel files
    for filename in sorted(os.listdir(XLSX_DIR)):
        if filename.endswith('.xlsx') and not filename.startswith('~'):
            filepath = os.path.join(XLSX_DIR, filename)
            info = extract_workbook_info(filepath)
            all_workbooks.append(info)

    # Save summary JSON
    summary_path = os.path.join(OUTPUT_DIR, 'workbooks_summary.json')
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(all_workbooks, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*60}")
    print(f"Summary saved to: {summary_path}")
    print(f"Total workbooks processed: {len(all_workbooks)}")

    # Print formula summary
    print(f"\n{'='*60}")
    print("FORMULA SUMMARY BY WORKBOOK")
    print('='*60)

    for wb in all_workbooks:
        total_formulas = sum(len(s['formulas']) for s in wb['sheets'])
        if total_formulas > 0:
            print(f"\n{wb['filename']} ({total_formulas} formulas):")
            for sheet in wb['sheets']:
                if sheet['formulas']:
                    print(f"  {sheet['name']}:")
                    for f in sheet['formulas'][:5]:  # Show first 5
                        print(f"    {f['cell']}: {f['formula']}")
                    if len(sheet['formulas']) > 5:
                        print(f"    ... and {len(sheet['formulas']) - 5} more")

if __name__ == '__main__':
    main()
